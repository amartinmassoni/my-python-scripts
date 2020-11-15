#!/usr/bin/python3

import openpyxl
import logging
import time
import re

class SmtpDelivery:
    def __init__( self, text, timestamp ):
        self.timestamp = timestamp
        self._to = None
        self._relay = None
        self._delay = None
        self._delays = None
        self._dsn = None
        self._status = None
        for item in text.split( ", " ):
            itemname, itemvalue = item.split( "=", 1 )
            setattr( self, "_" + itemname.lower(), itemvalue )

    def __repr__( self ):
        return f'<SmtpDelivery to: {self._to}, status: {self._status}>'

class QueueMessage:
    def __init__( self, queue_id, timestamp ):
        self.queue_id = queue_id
        self.from_timestamp = timestamp
        self.to_timestamp = timestamp
        self._uid = None
        self._client = None
        self._message_id = None
        self._from = None
        self._size = None
        self._nrcpt = None
        self.removed = False
        self.delivery = []
        self.items = []

    def add_line( self, timestamp, text ):
        self.to_timestamp = max( self.to_timestamp, timestamp )
        if text == "removed":
            self.removed = True
        elif text.startswith( "to=" ):
            # SmtpDelivery
            self.delivery.append( SmtpDelivery( timestamp = timestamp, text = text ) )
        else:
            for item in text.split( ", " ):
                itemname, itemvalue = item.split( "=", 1 )
                setattr( self, "_" + itemname.replace( "-", "_" ), itemvalue )

    def __repr__( self ):
        return f'<QueueMessage {self.queue_id}: from {self.from_timestamp} to {self.to_timestamp} removed:{self.removed}>'


def postfix_maillog( maillog ):
    preffix = "^([A-Z][a-z][a-z] [0-9 ][0-9] [0-9]{2}:[0-9]{2}:[0-9]{2}) [^ ]+ postfix/[a-z]+\[[0-9]+\]\: "
    re_message = re.compile( preffix + "([0-9A-F]{10,12}): (.+)$" )
    re_connect = re.compile( preffix + "connect from (.+)$" )
    re_disconnect = re.compile( preffix + "disconnect from (.+)$" )
    re_statistics = re.compile( preffix + "statistics: (.+)$" )
    re_timeout = re.compile( preffix + "(timeout after .+)$" )
    re_daemon = re.compile( preffix + "daemon started -- (.+)$" )
    current_year = time.strftime( "%Y" ) + " "
    queue_ids = []
    queue_messages = {}
    for filename in sorted( maillog ):
        logging.debug( "Reading file %s...", filename )
        for line in open( filename ):
            r1 = re_message.search( line )
            if r1:
                timestamp = time.strptime( current_year + r1.group( 1 ), "%Y %b %d %H:%M:%S" )
                queue_id = r1.group( 2 )
                if queue_id not in queue_messages:
                    queue_ids.append( queue_id )
                    queue_messages[ queue_id ] = QueueMessage( queue_id, timestamp )
                queue_messages[ queue_id ].add_line( timestamp, r1.group( 3 ) )
            elif re_connect.search( line ):
                pass
            elif re_disconnect.search( line ):
                pass
            elif re_statistics.search( line ):
                pass
            elif re_timeout.search( line ):
                pass
            elif re_daemon.search( line ):
                pass
            else:
                print( line )
    return ( queue_ids, queue_messages )

def messages_to_excel( queue_ids, queue_messages, output ):
    logging.debug( "Saving to %s...", output )
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Maillog"
    row_number = 1
    for ( column_number, column_name ) in enumerate( [ "Queue Id", "Uid", "Client", "Message-Id", "From", "Size", "Removed", "Delivery To", "Delivery Status" ], start = 1 ):
        ws1.cell( column = column_number, row = row_number, value = column_name )
    for queue_id in queue_ids:
        row_number = row_number + 1
        ws1.cell( column = 1, row = row_number, value = queue_id )
        queue_message = queue_messages[ queue_id ]
        ws1.cell( column = 2, row = row_number, value = queue_message._uid )
        ws1.cell( column = 3, row = row_number, value = queue_message._client )
        ws1.cell( column = 4, row = row_number, value = queue_message._message_id )
        ws1.cell( column = 5, row = row_number, value = queue_message._from )
        ws1.cell( column = 6, row = row_number, value = int( queue_message._size ) )
        ws1.cell( column = 7, row = row_number, value = queue_message.removed and "True" or "False" )
        if queue_message.delivery:
            row_number = row_number - 1
        for delivery in queue_message.delivery:
            row_number = row_number + 1
            ws1.cell( column = 8, row = row_number, value = delivery._to )
            ws1.cell( column = 9, row = row_number, value = delivery._status )
    wb.save( filename = output )

if __name__ == "__main__":

    import argparse

    parser = argparse.ArgumentParser()
    parser.description = "Process one or more postfix maillog files, and return one Excel file"
    parser.add_argument( "maillog", nargs = "+", help = "Input: maillog files" )
    parser.add_argument( "-o", "--output", help = "Output: Excel file", required = True )
    args = parser.parse_args()
    queue_ids, queue_messages = postfix_maillog( maillog = args.maillog )
    messages_to_excel( queue_ids = queue_ids, queue_messages = queue_messages, output = args.output )
